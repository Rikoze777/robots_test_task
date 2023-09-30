import openpyxl
from robots.services import fetch_robots
from openpyxl.styles import Alignment


def create_excel():
    week_robots = fetch_robots()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    workbook.remove_sheet(worksheet)
    page = 1

    for model, _ in week_robots.items():
        worksheet = workbook.create_sheet(model, page)
        worksheet.cell(row=1, column=1, value='Модель')
        worksheet.cell(row=1, column=2, value='Версия')
        worksheet.cell(row=1, column=3, value='Количество за неделю')

        alignment = Alignment(horizontal='center', vertical='center')

        column_letter = openpyxl.utils.get_column_letter(3)
        max_length = max(len(str(cell.value)) for cell
                         in worksheet[column_letter])
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

        page += 1
        row_num = 2

        for version, count in week_robots[model].items():
            worksheet.cell(row=row_num, column=1,
                           value=model)
            worksheet.cell(row=row_num, column=2,
                           value=version)
            worksheet.cell(row=row_num, column=3,
                           value=count)
            row_num += 1

        for row in worksheet.iter_rows(min_row=1,
                                       max_row=worksheet.max_row,
                                       min_col=1,
                                       max_col=3):
            for cell in row:
                cell.alignment = alignment

    return workbook
